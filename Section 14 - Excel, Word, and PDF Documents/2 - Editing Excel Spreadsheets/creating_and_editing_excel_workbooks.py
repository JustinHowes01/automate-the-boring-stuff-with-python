import openpyxl

wb = openpyxl.Workbook()

print(wb.sheetnames, end='\n\n')

sheet = wb['Sheet']

sheet['A1'] = 'Hello'
sheet['A2'] = 42

sheet2 = wb.create_sheet(title='New Sheet')
print(wb.sheetnames, end='\n\n')

wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames, end='\n\n')

wb.save('example.xlsx')
