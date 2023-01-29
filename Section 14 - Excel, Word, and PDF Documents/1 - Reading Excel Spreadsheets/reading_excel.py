import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')

sheet = workbook['Sheet1']

for i in range(1, 8):
    print(sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=2).value)
